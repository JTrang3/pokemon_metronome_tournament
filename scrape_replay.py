import os, re, ast
import openpyxl
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import filedialog

# Select replay(s) to extract data from
def select_replays():
    # Get the current working directory
    current_directory = os.getcwd()
    # Specify the folder name within the current directory
    folder_name = "replays"
    # Create the Tkinter root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    # Open the file dialog and allow the user to select multiple files
    file_paths = filedialog.askopenfilenames(initialdir=os.path.join(current_directory, folder_name), title="Select which replays to extract data from:")
    # Convert the returned tuple of file paths to a list
    selected_files = list(file_paths)
    # Convert absolute paths to relative paths
    selected_files = [os.path.relpath(file, current_directory) for file in selected_files]
    # Close the Tkinter root window (optional)
    root.destroy()
    return selected_files

# Convert html replay file into readable textfile
def convert_html_to_text(logs):
    # Get the name of the replay without the paths or extension
    base_name = os.path.splitext(os.path.basename(logs))[0]
    textfile = "logs\\" + base_name + ".txt"
    # Check if textfile already exists before converting
    if os.path.exists(textfile):
        return textfile
    else:
        with open(logs, 'r') as file:
            html_content = file.read()
        # Create a BeautifulSoup object
        soup = BeautifulSoup(html_content, "html.parser")
        # Process battle logs from elements in "inner message-log" div and store the lines
        message_log = soup.find("div", class_="inner message-log")
        elements = message_log.find_all(['div', 'h2'])
        lines = ''
        for element in elements:
            # Adjust team & roster format
            if element.name == 'div' and element.get('class') == ['chat', 'battle-history']:
                # Find the "strong" and "em" tags within the sub-div to add newlines
                strong_tag = element.find('strong')
                em_tag = element.find('em')
                if strong_tag:
                    lines += (strong_tag.get_text(strip=False) + '\n')
                if em_tag:
                    lines += (em_tag.get_text(strip=False) + '\n')
            # Adjust elements that print more than one line
            elif element.name == 'div' and element.get('class') == ['battle-history']:
                # Find all "break" tags within the sub-div
                break_tags = element.find_all('br')
                if len(break_tags) > 1:
                    # Add newline for each additional line printed
                    for break_tag in break_tags[:-1]:
                        break_tag.insert_after('\n')
                lines += (element.get_text(strip=False) + '\n')
            else:
                # Add newline to the text of the element
                if "DEBUG" not in element.get_text():
                    lines += (element.get_text(strip=False) + '\n')
        # Remove empty line at the end of the file
        lines = lines[:-1]
        # Write the lines as a new text file into logs folder
        with open(textfile, 'w') as file:
            file.writelines(lines)
        return textfile
    
# Gather all the moves used by Metronome in the entire battle
def metronome_data(logs):
    # {"move": "# of times used"}
    moves_used = {}
    # Open the file in read mode
    with open(logs, "r") as file:
        # Read the entire content of the file
        file_content = file.read()
        # Search for move occurrences using a pattern
        pattern = r"Waggling a finger let it use (.*?)!"
        move_matches = re.findall(pattern, file_content)
        # Count the moves and store in the dictionary
        for move in move_matches:
            moves_used[move] = moves_used.get(move, 0) + 1
    return moves_used

# Send move data to Excel spreadsheet
def transfer_metronome_data(dict, xl_file, sheetname):
    # Load the existing workbook & worksheet
    workbook = openpyxl.load_workbook(xl_file)
    worksheet = workbook[sheetname]
    # Iterate over each move and its occurences in the dictionary
    for key, value in dict.items():
        # Flag to check if the move found in the worksheet
        move_found = False  
        # Search for the move in the existing data
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == key:
                # Move already exists, increase the usage count
                move_found = True
                worksheet.cell(row=row_index, column=2).value += value
                break
        if move_found == False:
            # Move doesn't exist, add it to a new row
            new_row = worksheet.max_row + 1
            worksheet[f"A{new_row}"] = key
            worksheet[f"B{new_row}"] = value
    # Save the workbook
    workbook.save(xl_file)
    return

# Transfer move's stats from move datafile
def connect_move_with_stats(data_file, data_sheetname, target_file, target_sheetname):
    # Load worksheet where move database is
    stats_workbook = openpyxl.load_workbook(data_file)
    stats_worksheet = stats_workbook[data_sheetname]
    # Load worksheet where move data is going to
    target_workbook = openpyxl.load_workbook(target_file)
    target_worksheet = target_workbook[target_sheetname]
    # Iterate over the rows in the target worksheet
    for row_index, row in enumerate(target_worksheet.iter_rows(min_row=2, values_only=True), start=2):
        # Flag to check for valid move
        valid_move = False
        # Check if the move doesn't contains its stats (which start in column C)
        if target_worksheet.cell(row=row_index, column=3).value is None:
            move_name, count = row[0], row[1]
            # Search for the move in the stats worksheet (moves start in column B)
            for stats_row in stats_worksheet.iter_rows(min_row=2, min_col=2, values_only=True):
                if stats_row[0] == move_name:
                    valid_move = True
                    # Retrieve Type, Category, Power, Accuracy, Generation, Effect stats (columns C-F, H-I)
                    move_stats = stats_row[1:5] + stats_row[6:8]
                    # Update the corresponding cells in the target worksheet
                    target_worksheet.cell(row=row_index, column=2).value = count
                    # Add stats starting from column C
                    for col_index, stat in enumerate(move_stats, start=3):
                        target_worksheet.cell(row=row_index, column=col_index).value = stat
                    # Once move is found & stats populated, stop looping through datasheet
                    break
            if valid_move == False:
                # Not a valid move, remove from data
                target_worksheet.delete_rows(row_index)
                print(f"Invalid move: {move_name} removed")
    # Save the target workbook
    target_workbook.save(target_file)
    return

# Initialize dictionary w/ pokemon's team & match data
def get_teams_and_roster(logs):
    with open(logs, 'r') as file:
        # Read lines 3-6 in battle logs for teams
        file_content = "".join(file.readlines()[2:6])
        # Search for team type and rosters using a pattern
        pattern = r"MTN\s+(\w+)'s team:\n([\w\s/-]+)\n"
        teams = re.findall(pattern, file_content)
        return teams

# Returns the team that won the battle
def get_victor(logs):
    with open(logs, 'r') as file:
        lines = file.readlines()
        last_line = lines[-1].strip()
        team_won = re.search(r"MTN\s+(\w+) won the battle!", last_line).group(1)
    return team_won

# Read through pokemon database & get pokemon's type
def get_pokemon_type(pokemon, xl_file, sheetname):
    # Load file where pokemon data is found
    workbook = openpyxl.load_workbook(xl_file)
    worksheet = workbook[sheetname]
    # Search for pokemon in database
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[1] == pokemon:
            # Pokemon found, combine pokemon's types into a list
            types = [worksheet.cell(row=row_index, column=col_index).value for col_index in range(3,5) if worksheet.cell(row=row_index, column=col_index).value != None]
            return types
    # Pokemon not found
    print("Pokemon not found")

# Determine which method a pokemon faints
def faint_case(pokemon, case_num):
    match case_num:
        case "STATUS":
            print(f"{pokemon}: Status")
        case "CONFUSION":
            print(f"{pokemon}: Confusion")
        case "SELF-DESTRUCT":
            print(f"{pokemon}: Recoil/Self-destruct")
        case "CONTACT":
            print(f"{pokemon}: Contact-damage")
        case "TK":
            print(f"{pokemon}: Team kill")
        case "FUTURE":
            print(f"{pokemon}: Future Sight/Doom Desire")
        case "PERISH (SELF)":
            print(f"{pokemon}: Perish Song (self-destruct)")
        case "PERISH (TK)":
            print(f"{pokemon}: Perish Song (team kill)")
        case "PERISH":
            print(f"{pokemon}: Perish Song")
        case "WEATHER":
            print(f"{pokemon}: Hail/Sandstorm")
        case "HAZARDS":
            print(f"{pokemon}: Stealth Rock/Spikes")
        case "DESTINY":
            print(f"{pokemon}: Destiny Bond")
        case "NIGHTMARE":
            print(f"{pokemon}: Nightmare")
        case "CURSE":
            print(f"{pokemon}: Curse")
        case "BIND":
            print(f"{pokemon}: Bind")
        case "LEECH":
            print(f"{pokemon}: Leech Seed")
        case _:
            print(f"{pokemon}: Normal")

# Gather battle statistics from the entire battle
turn_count = 0
def match_data(logs):
    # Scoreboard: {"pokemon": ["team type", "defeats", "faints", "self-faints", "damage given", "damage taken", "battle turns", "games played", {"affliction": "pokemon"}, "turn_active"]}
    match_scoreboard = {}
    # Get teams and their rosters
    teams = get_teams_and_roster(logs)
    for team in teams:
        # Separate team type and roster
        team_type = team[0]
        roster = team[1].split(" / ")
        for pokemon in roster:
            # Apply initial scoreboard to each pokemon in roster
            match_scoreboard[pokemon.split('-')[0]] = [team_type, 0, 0, 0, 0, 0, 0, 1, {}, False] 
    global turn_count
    # Read the contents of the file
    with open(logs, 'r') as file:
        next_line = file.readlines()
    with open(logs, 'r') as file:
        future_attacking, future_defeat_score = [], None
        perish_user, perish_faints, perish_counted = [], 0, False
        line_index = 1
        case_num = ""
        future_outcomes = ["avoided the attack!", "It doesn't affect", "(Future Sight did not hit because the target is fainted.)", "(Doom Desire did not hit because the target is fainted.)"]
        bind_moves = ["was squeezed by", "was wrapped by", "clamped down on", "became trapped in", "became trapped by"]
        weather, weather_user = ["A sandstorm kicked up!", "It started to hail!", "The sunlight turned harsh!", "It started to rain!"], ""
        # Apply initial hazards for each team
        opponent_type = [team[0] for team in teams]
        hazards = {opponent_type[0]: {}, opponent_type[1]: {}}
        for line in file:
            # Set all active pokemon for current turn
            if line_index >= 8:
                for pokemon in match_scoreboard:
                    if pokemon in line:
                        match_scoreboard[pokemon][9] = True
            # Check who last used Metronome
            if " used " in line:
                current_action = re.search(r"(?:The opposing )?([^ ]+) used ([^!]+)!", line)
                pokemon_attacking = current_action.group(1)
                move_used = current_action.group(2)
                # print(f"Pokemon: {pokemon_attacking}, Move: {current_action}")
            # Pokemon uses Future Sight/Doom Desire
            elif "foresaw an attack!" in line or "chose Doom Desire as its destiny!" in line:
                # Store future attacks in a list to check precedence
                future_attacking.append(pokemon_attacking)
            # Future Sight/Doom Desire misses the target
            elif "took the Future Sight attack!" in line or "took the Doom Desire attack!" in line:
                if any(future_message in next_line[line_index] for future_message in future_outcomes):
                    future_attacking.pop(0)
            # Pokemon uses Perish Song
            elif "that heard the song will faint in three turns!" in line:
                # Store Perish Song users in a list to check precedence
                perish_user.append(pokemon_attacking)
            # Gather all pokemon fainted by Perish Song
            elif "perish count fell to 0." in line:
                perish_faints += 1
            # Pokemon affected by status condition
            elif "was poisoned!" in line or "was badly poisoned!" in line or "was burned!" in line:
                ability_status = re.search(r"(?<=\[)(?:The opposing )?([^ ]+)'s ([^\]]+)", previous_line)
                pokemon_status = re.search("(?:The opposing )?([^ ]+).*!", line).group(1)
                if "'s " in previous_line:
                    # Attacking pokemon statused by receiver's ability
                    if ability_status.group(1) != pokemon_attacking:
                        match_scoreboard[pokemon_attacking][8]["Status"] = ability_status.group(1)
                    # Receiving pokemon statused by attacker's ability
                    else:
                        match_scoreboard[pokemon_receiving][8]["Status"] = ability_status.group(1)
                # Receiving pokemon statused by hazards
                elif "Go!" in previous_line or "sent out" in previous_line or "was hurt by the spikes!" in previous_line or "Pointed stones dug into" in previous_line:
                    match_scoreboard[pokemon_status][8]["Status"] = hazards[match_scoreboard[pokemon_status][0]]["T-Spikes"]
                # Receiving pokemon statused by attacker's move
                else:
                    match_scoreboard[pokemon_status][8]["Status"] = pokemon_attacking
            # Pokemon affected by Leech Seed
            elif "was seeded!" in line:
                # Update receiving pokemon's scoreboard with "leech" affliction
                pokemon_leech = re.search(r"(?:The opposing )?([^ ]+) was seeded!", line).group(1)
                match_scoreboard[pokemon_leech][8]["Leech"] = pokemon_attacking
            # Pokemon sets up hazards
            elif "Pointed stones float in the air around" in line or "scattered on the ground all around" in line:
                # Hazards on user's side of the field set by opposing team type
                receiving_team = opponent_type[1] if match_scoreboard[pokemon_attacking][0] == opponent_type[0] else opponent_type[0]
                if "Pointed stones float in the air around" in line:
                    hazards[receiving_team]["Rocks"] = pokemon_attacking
                elif "Spikes were scattered on the ground all around" in line:
                    hazards[receiving_team]["Spikes"] = pokemon_attacking
                else:
                    hazards[receiving_team]["T-Spikes"] = pokemon_attacking
            # Pokemon hit with binding move
            elif any(bind_message in line for bind_message in bind_moves):
                # Binded pokemon's name is at the end with Clamp
                if "clamped down on" in line:
                    pokemon_receiving = re.search(r"(?:The opposing )?([^ ]+) clamped down on (?:the opposing )?([^ ]+)!", line).group(2)
                else:
                    pokemon_receiving = re.search(r"(?:The opposing )?([^ ]+).*!", line).group(1)
                # Update receiving pokemon's scoreboard with "bind" affliction
                match_scoreboard[pokemon_receiving][8]["Bind"] = pokemon_attacking
            # Pokemon hit with contact-damage ability
            elif "was hurt!" in line or "sucked up the liquid ooze!" in line:
                pokemon_contact = re.search(r"(?<=\[)(?:The opposing )?([^ ]+)'s ([^\]]+)", previous_line).group(1)
            # Pokemon hit with Nightmare
            elif "began having a nightmare!" in line:
                # Update receiving pokemon's scoreboard with "nightmare" affliction
                pokemon_nightmare = re.search(r"(?:The opposing )?([^ ]+) began having a nightmare!", line).group(1)
                match_scoreboard[pokemon_nightmare][8]["Nightmare"] = pokemon_attacking
            # Pokemon hit with Curse from Ghost type
            elif "cut its own HP and put a curse on" in line:
                # Update receiving pokemon's scoreboard with "curse" affliction
                pokemon_curse = re.search(r"(?:The opposing )?([^ ]+) cut its own HP and put a curse on (?:the opposing )?([^ ]+)!", line).group(2)
                match_scoreboard[pokemon_curse][8]["Curse"] = pokemon_attacking
            # Current state of the weather
            elif any(weather_message in line for weather_message in weather):
                # Pokemon used weather move
                if " use " in previous_line:
                    weather_user = pokemon_attacking
                # Pokemon used weather ability
                else:
                    weather_user = re.search(r"(?<=\[)(?:The opposing )?([^ ]+)'s ([^\]]+)", previous_line).group(1)
            # Check for damage dealt
            elif "of its health!" in line:
                pokemon_receiving = re.search(r"(?:The opposing )?[(]?([^ ]+) lost (\d+\.?\d*)% of its health!", line).group(1)
                damage = float(re.search(r"(?:The opposing )?[(]?([^ ]+) lost (\d+\.?\d*)% of its health!", line).group(2))
                # print(f"Pokemon: {pokemon_receiving}, Damage: {damage}")
                # Update receiving pokemon's "damage taken" & attacking pokemon's "damage given" stat
                match_scoreboard[pokemon_receiving][5] += damage
                # Check if pokemon hit by Future Sight/Doom Desire
                if "took the Future Sight attack!" in previous_line or "took the Doom Desire attack!" in previous_line:
                    # Take from list of future attacks & remove earliest attack
                    match_scoreboard[future_attacking[0]][4] += damage
                    if "fainted!" in next_line[line_index + 1]:
                        future_defeat_score = future_attacking[0]
                    future_attacking.pop(0)
                # NOTE: Pokemon won't receive "damage given" score if attacking own team
                elif match_scoreboard[pokemon_attacking][0] != match_scoreboard[pokemon_receiving][0]:
                    match_scoreboard[pokemon_attacking][4] += damage
            # Check for pokemon fainting
            elif "fainted!" in line:
                pokemon_fainted = re.search(r"(?:The opposing )?([^ ]+) fainted!", line).group(1)
                # Update fainted pokemon's "faints" stat
                match_scoreboard[pokemon_fainted][2] += 1
                # Check who receives "defeats" score for faint
                # CASE 1: Pokemon faints from Future Sight/Doom Desire
                if future_defeat_score != None:
                    match_scoreboard[future_defeat_score][1] += 1
                    future_defeat_score = None
                    case_num = "FUTURE"
                # CASE 2: Pokemon faints from Perish Song
                elif perish_faints != 0:
                    # CASE 2.1: Perish Song user faints itself
                    if perish_user[0] == pokemon_fainted:
                        match_scoreboard[pokemon_fainted][3] += 1
                        case_num = "PERISH (SELF)"
                    # CASE 2.2: Perish Song user faints team member
                    elif match_scoreboard[perish_user[0]][0] == match_scoreboard[pokemon_fainted][0]:
                        case_num = "PERISH (TK)"
                        pass
                    # Update perish user's "defeats" score
                    else:
                        match_scoreboard[perish_user[0]][1] += 1
                        case_num = "PERISH"
                    perish_faints -= 1
                    # Counted all perished pokemon in current turn
                    if perish_faints == 0:
                        perish_counted = True
                # CASE 3: Pokemon faints from binding move (Bind/Wrap/Fire Spin/etc.)
                elif "is hurt by" in previous_line:
                    # Give "defeats" score to binding pokemon
                    match_scoreboard[match_scoreboard[pokemon_fainted][8]["Bind"]][1] += 1
                    case_num = "BIND"
                # CASE 3.1: Pokemon faints from residual damage (poison/burn)
                elif "was hurt by poison!" in previous_line or "was hurt by its burn!" in previous_line:
                    # Give "defeats" score to pokemon who statused
                    match_scoreboard[match_scoreboard[pokemon_fainted][8]["Status"]][1] += 1
                    case_num = "STATUS"
                # CASE 3: Pokemon faints from Leech Seed
                elif "health is sapped by Leech Seed!" in previous_line:
                    match_scoreboard[match_scoreboard[pokemon_fainted][8]["Leech"]][1] += 1
                    case_num = "LEECH"
                # CASE 3.2: Pokemon faints from residual damage (sandstorm/hail)
                elif "is buffeted by the" in previous_line or "hurt by its Dry Skin" in previous_line:
                    # Pokemon faints from own weather
                    if pokemon_fainted == weather_user:
                        match_scoreboard[weather_user][3] += 1
                    # Pokemon faints from team's weather
                    elif match_scoreboard[pokemon_fainted][0] == match_scoreboard[weather_user][0]:
                        # Weather user won't receive "defeats" score
                        pass
                    else:
                        # Update weather user's "defeats" score
                        match_scoreboard[weather_user][1] += 1
                    case_num = "WEATHER"
                # CASE 3.3: Pokemon faints from residual damage (hazards)
                elif "Pointed stones dug into" in previous_line or "was hurt by the spikes!" in previous_line:
                    # Give "defeats" score to pokemon(s) who set up hazards
                    if "Pointed stones dug into" in previous_line:
                        match_scoreboard[hazards[match_scoreboard[pokemon_fainted][0]]["Rocks"]][1] += 1
                    else:
                        match_scoreboard[hazards[match_scoreboard[pokemon_fainted][0]]["Spikes"]][1] += 1
                    case_num = "HAZARDS"
                # CASE 4: Pokemon faints from Destiny Bond
                elif "took its attacker down with it!" in previous_line:
                    # Give "defeats" score to pokemon who used Destiny Bond
                    pokemon_bonded = re.search(r"(?:The opposing )?([^ ]+) took its attacker down with it!", previous_line).group(1)
                    match_scoreboard[pokemon_bonded][1] += 1
                    case_num = "DESTINY"
                # CASE 5: Pokemon faints from Nightmare
                elif "is locked in a nightmare!" in previous_line:
                    match_scoreboard[match_scoreboard[pokemon_fainted][8]["Nightmare"]][1] += 1
                    case_num = "NIGHTMARE"
                # CASE 6: Pokemon faints from Curse
                elif "is afflicted by the curse!" in previous_line:
                    match_scoreboard[match_scoreboard[pokemon_fainted][8]["Curse"]][1] += 1
                    case_num = "CURSE"
                # CASE 7.1: Pokemon faints from itself (confusion)
                elif "It hurt itself in its confusion!" in previous_line:
                    match_scoreboard[pokemon_fainted][3] += 1
                    case_num = "CONFUSION"
                # CASE 7.2: Pokemon faints from itself (recoil/self-destruct)
                elif pokemon_fainted == pokemon_attacking:
                    # CASE 7.3: Pokemon faints from contact-damage ability 
                    if "was hurt!" in previous_line or "sucked up the liquid ooze!" in previous_line:
                        # Give "defeats" score to pokemon with contact ability
                        match_scoreboard[pokemon_contact][1] += 1
                        case_num = "CONTACT" 
                    else:
                        match_scoreboard[pokemon_fainted][3] += 1
                        case_num = "SELF-DESTRUCT"
                # CASE 8: Pokemon faints from own team member
                elif match_scoreboard[pokemon_fainted][0] == match_scoreboard[pokemon_attacking][0]:
                    # Attacking pokemon won't receive "defeats" score
                    case_num = "TK"
                    pass
                # Update attacking pokemon's "defeats" stat
                else:
                    match_scoreboard[pokemon_attacking][1] += 1
                    case_num = 0
                # Testing proper faint cases
                faint_case(pokemon_fainted, case_num)
            elif "Turn " in line:
                # All pokemon perished from field, remove earliest perish user
                if perish_counted == True:
                    perish_user.pop(0)
                    perish_counted = False
                # Set current turn of battle and active turns of each pokemon
                turn_count += 1
                for pokemon in match_scoreboard:
                    if match_scoreboard[pokemon][9] == True:
                        match_scoreboard[pokemon][6] += 1
                        match_scoreboard[pokemon][9] = False
            # Ignore empty lines in file
            if line.strip() != "":
                previous_line = line
            if line_index < len(next_line):
                line_index += 1
    return match_scoreboard

# Send battle data statistics to Excel spreadsheet
def transfer_match_data(dict, xl_file, sheetname, pokemon_data):
    # Load the existing workbook & worksheet
    workbook = openpyxl.load_workbook(xl_file)
    worksheet = workbook[sheetname]
    # Iterate over the key, values of each pokemon in the scoreboard
    for pokemon, scoreboard in dict.items():
        # Flag to check if pokemon already in the worksheet
        pokemon_found = False
        # Search for the pokemon in the worksheet
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == pokemon:
                # Pokemon found, append the stats to the existing row
                pokemon_found = True
                # Ignore "team type", "affliction", & "turn_active" values in dict
                for col_index, stat in enumerate(scoreboard[:-2], start=3):
                    if col_index != 3:
                        worksheet.cell(row=row_index, column=col_index).value += stat
        if pokemon_found == False:
            # Pokemon not found, add the stats to a new row
            new_row = worksheet.max_row + 1
            worksheet.cell(row=new_row, column=1).value = pokemon
            # Call function to get pokemon's type
            type = get_pokemon_type(pokemon, pokemon_data, sheetname)
            worksheet.cell(row=new_row, column=2).value = str(type)
            # Ignore "affliction" & "turn_active" values in dict
            for col_index, stat in enumerate(scoreboard[:-2], start=3):
                worksheet.cell(row=new_row, column=col_index).value = stat
    # Save the workbook
    workbook.save(xl_file)
    return

# Gather team statistics from scoreboard
def team_data(logs, dict):
    # {"pokemon": ["team type", "defeats", "faints", "self-faints", "damage given", "damage taken", "battle turns", "games played"]}
    # {"team": [["members"], "won", "lost", "match turns", "games played", ["match history"]]}
    team_match_history = {} 
    # Get teams and their rosters
    teams = get_teams_and_roster(logs)
    winner = get_victor(logs)
    global turn_count
    # Create match statistics for each team 
    for team in teams:
        # Separate team type and roster
        team_type = team[0]
        roster = team[1].split(" / ")
        # Set winner & loser
        win, lose = (1, 0) if team_type == winner else (0, 1)
        # Get total faints for each team & opponent team type from scoreboard
        user_faints, opponent_faints, opponent_type = 0, 0, ""
        for stat in dict.values():
            if stat[0] == team_type:
                # Sum ally team's faints
                user_faints += stat[2]
            elif stat[0] != team_type: 
                # Sum enemy team's faints
                opponent_faints += stat[2]
                opponent_type = stat[0]
            # Populate team scoreboard with match records
            team_match_history[team_type] = [roster, win, lose, turn_count, 1, [f"{opponent_type}: {opponent_faints}-{user_faints}"]]
    # Reset turn count for next match
    turn_count = 0
    return team_match_history

# Send team records to Excel spreadsheet
def transfer_team_data(dict, xl_file, sheetname):
    # Load the existing workbook & worksheet
    workbook = openpyxl.load_workbook(xl_file)
    worksheet = workbook[sheetname]
    # Iterate over both teams in the dictionary
    for team, records in dict.items():
        # Flag to check if team is already in worksheet
        team_found = False
        # Search for the team in the worksheet
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == team:
                # Team found, append records to the existing row
                team_found = True
                for col_index, stat in enumerate(records, start=2):
                    # Ignore "members" value in dict
                    if col_index >= 3:
                        if isinstance(stat, int):
                            worksheet.cell(row=row_index, column=col_index).value += stat
                        else:
                            # Format match history into proper list using abstract syntax trees
                            match_history = ast.literal_eval(worksheet.cell(row=row_index, column=col_index).value)
                            match_history += stat
                            worksheet.cell(row=row_index, column=col_index).value = str(match_history)
        if team_found == False:
            # Team not found, add records to a new row
            new_row = worksheet.max_row + 1
            worksheet.cell(row=new_row, column=1).value = team
            for col_index, stat in enumerate(records, start=2):
                if isinstance(stat, int):
                    worksheet.cell(row=new_row, column=col_index).value = stat
                else:
                    worksheet.cell(row=new_row, column=col_index).value = str(stat)
    # Save the workbook
    workbook.save(xl_file)
    return
    
def main():
    logs = []
    user_replays = select_replays()

    for replay in user_replays:
        replay_log = convert_html_to_text(replay)
        logs.append(replay_log)

    for log in logs:
        print(f"\n{log}")

        moves = metronome_data(log)
        scoreboard = match_data(log)
        team_record = team_data(log, scoreboard)

        # TEST: View complete scoreboard
        for key, values in scoreboard.items():
            print(f"{key}: {values}")
        # TEST: View complete match records
        for key, values in team_record.items():
            print(f"{key}: {values}")

        transfer_metronome_data(moves, "tournament_statistics.xlsx", "Metronome_Data")
        connect_move_with_stats("pokemon_data_gen5.xlsx", "Move_Data", "tournament_statistics.xlsx", "Metronome_Data")
        transfer_match_data(scoreboard, "tournament_statistics.xlsx", "Pokemon_Data", "pokemon_data_gen5.xlsx")
        transfer_team_data(team_record, "tournament_statistics.xlsx", "Team_Data")
    
    # TEST: Count each move used & total moves used in a battle
    # for move, count in moves.items():
    #     print(f"{move}: {count}")
    # print(f"Keys: {len(moves.keys())}, Values: {sum(moves.values())}")

    # TEST: Check how moves are being transferred between files
    # test_dict = {"Pound": 1, "Fire Punch": 2, "Cut": 1, "Swords Dance": 1, "Scratch": 2, "Fusion Bolt": 1, "Fly": 1}

    # TEST: View initial team rosters
    # print(get_teams_and_roster(battle_log))

    # TEST: View complete scoreboard
    # for key, values in scoreboard.items():
    #     print(f"{key}: {values}")

    # TEST: View complete match records
    # for key, values in team_record.items():
    #     print(f"{key}: {values}")

if __name__ == "__main__":
    main()